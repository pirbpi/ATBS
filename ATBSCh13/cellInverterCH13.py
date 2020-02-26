#! python3
# cellInverter.py inverts the row and column of the cells in the spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
output_wb = openpyxl.Workbook()
invertedSheet = output_wb.active
rowMatrix = []  # Create an empty list

for row in range(sheet.max_row):  # For 'row' in range 0 to the last row w/ dat
    rowMatrix.append([])  # Create an empty sublist inside the list
    for cellObj in list(sheet.rows)[row]:  # For each cell in the current row
        rowMatrix[row].append(cellObj.value)  # Add the cell to the sub-list

for col in range(len(rowMatrix)):
    for i in range(1, len(rowMatrix[0])+1):
        invertedSheet.cell(row=i, column=1).value = rowMatrix[0][i-1]
    for row in range(len(rowMatrix[row])):
        invertedSheet.cell(row=row+1, column=col+1).value = rowMatrix[col][row]

output_wb.save('produceEdited.xlsx')

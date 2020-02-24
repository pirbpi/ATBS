#! python3
# multiplicationTable - take a cmd line argument and create a multiplication table

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

userMultiplication = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active


for i in range(1, userMultiplication+1):
    sheet.cell(row=1,column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = Font(bold=True)
    sheet.cell(row=i + 1, column = 1).value = i
    sheet.cell(row=i + 1, column=1).font = Font(bold=True)

for col in range(2, userMultiplication+2):
    for row in range(2, userMultiplication+2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col - 1) * (row - 1)

wb.save('multiplicationTable.xlsx')
